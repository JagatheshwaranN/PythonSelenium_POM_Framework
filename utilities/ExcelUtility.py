"""
Name : ExcelUtility.py
Author : Jaga
Date : 21-02-2021
"""

import openpyxl


class ExcelReader:

    @staticmethod
    def get_row_count(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheet_name)
        return sheet.max_row

    @staticmethod
    def get_column_count(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheet_name)
        return sheet.max_column

    @staticmethod
    def read_data(file, sheet_name, row_num, column_num):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheet_name)
        return sheet.cell(row=row_num, column=column_num).value

    @staticmethod
    def write_data(file, sheet_name, row_num, column_num, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheet_name)
        sheet.cell(row=row_num, column=column_num).value = data
        workbook.save(file)
